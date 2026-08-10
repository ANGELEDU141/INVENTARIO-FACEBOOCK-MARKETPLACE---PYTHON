# -*- coding: utf-8 -*-
"""
FACEBOOK MARKETPLACE - EXTRACTOR FINAL

Basado en el código original entregado por el usuario.

Cambios principales:
- Mantiene el detector de tarjetas original (img[alt] + ancestros).
- El NOMBRE SIEMPRE se toma primero del alt de la imagen de la tarjeta.
  Esto evita que "Tu publicación" termine como nombre.
- Precio: S/285S/305 -> 285.
- SKU: se obtiene del modal mediante "3023 / Publicación:".
- Disponibilidad y fecha: se obtienen del modal.
- listing_id: se intenta obtener del modal/HTML.
- Salida FINAL en Excel (.xlsx), no CSV.
- SKU y listing_id se escriben como texto para evitar notación científica.
- Conserva registros anteriores desde el Excel existente.
- Si existe el CSV viejo, lo importa automáticamente una sola vez.
- Guarda el Excel después de cada publicación.
"""

import csv
import json
import re
import time
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from playwright.sync_api import sync_playwright

CDP_URL = "http://127.0.0.1:9222"
MARKETPLACE_URL = "https://www.facebook.com/marketplace/you/selling"

BASE = Path(__file__).resolve().parent

EXCEL_SALIDA = BASE / "Facebook_Marketplace.xlsx"
CSV_VIEJO = BASE / "Facebook_Marketplace.csv"

PAUSA_ENTRE_PRODUCTOS = 0.35
PAUSA_SCROLL = 0.45
MAX_SCROLLS = 80
TIMEOUT_MODAL = 5000

CAMPOS = [
    "nombre",
    "precio",
    "disponibilidad",
    "sku",
    "publicado",
    "listing_id",
]


# ============================================================
# TEXTO
# ============================================================

def limpiar_texto(texto):
    if not texto:
        return ""

    texto = texto.replace("\xa0", " ")

    return "\n".join(
        re.sub(r"\s+", " ", x).strip()
        for x in texto.replace("\r", "\n").split("\n")
        if x.strip()
    )


def limpiar_nombre(nombre):
    """Limpieza segura de nombres provenientes de Facebook/Excel."""
    if nombre is None:
        return ""
    n = str(nombre).replace("\xa0", " ")
    n = n.replace("\r", " ").replace("\n", " ")
    n = re.sub(r"\s+", " ", n).strip()
    # Solo elimina comillas exteriores agregadas por CSV/Excel.
    # Conserva las internas: 15.6", 22", 2.5", etc.
    if len(n) >= 2 and n[0] == '"' and n[-1] == '"':
        n = n[1:-1].strip()
    # Corrige el caso heredado 15.6"" -> 15.6".
    n = re.sub(r'(?<=\d)""(?=\s|[A-Za-zÁÉÍÓÚÑáéíóúñ]|$)', '"', n)
    return n.strip()


def normalizar(nombre):
    return re.sub(r"\s+", " ", limpiar_nombre(nombre)).strip().lower()


def registro_key(registro):
    """Identidad: listing_id primero; nombre solo como fallback sin ID."""
    listing = str(registro.get("listing_id") or "").strip()
    if listing:
        return "id:" + listing
    nombre = normalizar(registro.get("nombre", ""))
    return "name:" + nombre if nombre else ""


def fusionar_registro(viejo, nuevo):
    """Actualiza datos nuevos sin borrar datos existentes no vacíos."""
    resultado = registro_vacio()
    for campo in CAMPOS:
        a = str((viejo or {}).get(campo, "") or "").strip()
        b = str((nuevo or {}).get(campo, "") or "").strip()
        resultado[campo] = b if b else a
    resultado["nombre"] = limpiar_nombre(resultado["nombre"])
    return resultado



# ============================================================
# PRECIO / ESTADO / PUBLICADO / SKU
# ============================================================

def precios(texto):
    """
    Devuelve todos los precios encontrados.

    S/285S/305 -> ["285", "305"]
    """
    return [
        x.replace(",", "")
        for x in re.findall(
            r"S/\s*([\d.,]+)",
            texto or "",
            re.I
        )
    ]


def precio(texto):
    p = precios(texto)
    return p[0] if p else ""


def disponibilidad(texto):
    """Obtiene el estado real evitando falsos positivos de los botones."""
    lineas = [x.strip() for x in (texto or "").splitlines() if x.strip()]
    # La línea del precio manda: S/5309 · Agotado · Miraflores, LM
    for linea in lineas:
        if re.search(r"S/\s*[\d.,]+", linea, re.I):
            if re.search(r"\bAgotado\b", linea, re.I):
                return "Agotado"
            if re.search(r"\bDisponible\b", linea, re.I):
                return "Disponible"
    # Fallback, ignorando botones de acción.
    for linea in lineas:
        if re.search(r"Marcar como (agotado|disponible)", linea, re.I):
            continue
        if re.search(r"\bAgotado\b", linea, re.I):
            return "Agotado"
        if re.search(r"\bDisponible\b", linea, re.I):
            return "Disponible"
    return ""


def publicado(texto):
    """Devuelve cuándo se publicó, no dónde está publicada."""
    lineas = [x.strip() for x in (texto or "").splitlines() if x.strip()]
    for i, linea in enumerate(lineas):
        m = re.search(r"Publicaci[oó]n\s*:\s*(.+)$", linea, re.I)
        if m:
            return m.group(1).strip()
        if re.fullmatch(r"Publicaci[oó]n\s*:?", linea, re.I) and i + 1 < len(lineas):
            return lineas[i + 1].strip()
    for linea in lineas:
        if re.fullmatch(r"Hace\s+.+", linea, re.I):
            return linea
    return ""


def sku_desde_modal(texto):
    """
    Facebook:

    3023
    Publicación:
    Hace 18 horas
    """

    lineas = [
        x.strip()
        for x in (texto or "").splitlines()
        if x.strip()
    ]

    for i, linea in enumerate(lineas):

        if re.fullmatch(
            r"Publicaci[oó]n\s*:?",
            linea,
            re.I
        ) and i > 0:

            anterior = lineas[i - 1]

            if re.fullmatch(
                r"\d+",
                anterior
            ):
                return anterior

    m = re.search(
        r"(\d+)\s*Publicaci[oó]n\s*:",
        texto or "",
        re.I
    )

    return m.group(1) if m else ""


# ============================================================
# NOMBRE
# ============================================================

def nombre_valido(nombre):
    """
    Evita que textos del modal terminen como nombre.
    """

    if not nombre:
        return False

    n = nombre.strip()

    if len(n) < 3:
        return False

    prohibidos_exactos = {
        "tu publicación",
        "tu publicacion",
        "publicación",
        "publicacion",
        "marketplace",
        "facebook",
        "imagen",
        "image",
    }

    if n.lower() in prohibidos_exactos:
        return False

    if re.fullmatch(
        r"S/\s*[\d.,]+(?:S/\s*[\d.,]+)*",
        n,
        re.I
    ):
        return False

    if re.fullmatch(
        r"\d+",
        n
    ):
        return False

    return True


def nombre_desde_texto(texto):
    """
    Fallback para el modal.

    MUY IMPORTANTE:
    "Tu publicación" está explícitamente excluido.
    """

    lineas = [
        x.strip()
        for x in (texto or "").splitlines()
        if x.strip()
    ]

    for linea in lineas:

        if not nombre_valido(linea):
            continue

        if re.search(
            r"^(Disponible|Agotado)\b",
            linea,
            re.I
        ):
            continue

        if re.search(
            r"Publicado\s+(el|hace|en)",
            linea,
            re.I
        ):
            continue

        if re.search(
            r"Publicado en Marketplace",
            linea,
            re.I
        ):
            continue

        if re.search(
            r"clics?\s+en\s+la publicación",
            linea,
            re.I
        ):
            continue

        if re.search(
            r"Marcar como agotado|Promocionar publicación|Editar publicación|Eliminar publicación",
            linea,
            re.I
        ):
            continue

        return linea

    return ""


def nombre_desde_tarjeta(img, texto):
    """
    PRIORIDAD:

    1. img[alt] real de la tarjeta.
    2. aria-label del contenedor.
    3. texto de la tarjeta.

    Nunca devuelve "Tu publicación".
    """

    try:
        alt = (
            img.get_attribute("alt")
            or ""
        ).strip()

        if nombre_valido(alt):
            return alt
    except Exception:
        pass

    return nombre_desde_texto(texto)


# ============================================================
# LISTING ID
# ============================================================

def listing_id_desde_html(html):
    if not html:
        return ""

    patrones = [
        r'"listing_id"\s*:\s*"(\d{5,})"',
        r'"listing_id"\s*:\s*(\d{5,})',
        r'listing_id=(\d{5,})',
        r'listing_id%3D(\d{5,})',
    ]

    for patron in patrones:

        m = re.search(
            patron,
            html,
            re.I
        )

        if m:
            return str(
                m.group(1)
            )

    return ""


def listing_id_modal(modal):
    """
    Primero busca enlaces reales.
    Después HTML del modal.
    """

    try:

        links = modal.locator(
            '[href*="listing_id="]'
        )

        for i in range(
            links.count()
        ):

            href = (
                links.nth(i)
                .get_attribute("href")
                or ""
            )

            m = re.search(
                r"listing_id(?:=|%3D)(\d+)",
                href,
                re.I
            )

            if m:
                return m.group(1)

    except Exception:
        pass

    try:

        html = modal.inner_html()

        resultado = listing_id_desde_html(
            html
        )

        if resultado:
            return resultado

    except Exception:
        pass

    return ""


# ============================================================
# EXCEL
# ============================================================

def registro_vacio():
    return {
        "nombre": "",
        "precio": "",
        "disponibilidad": "",
        "sku": "",
        "publicado": "",
        "listing_id": "",
    }


def cargar_excel():
    """Carga Excel, limpia nombres y elimina encabezados repetidos."""
    registros = {}
    if not EXCEL_SALIDA.exists():
        return registros
    try:
        wb = load_workbook(EXCEL_SALIDA, read_only=True, data_only=True)
        ws = wb.active
        encabezados = [str(c.value).strip().lower() if c.value is not None else "" for c in ws[1]]
        indices = {nombre: i for i, nombre in enumerate(encabezados)}
        if "nombre" not in indices:
            wb.close()
            return registros
        for fila in ws.iter_rows(min_row=2, values_only=True):
            registro = registro_vacio()
            for campo in CAMPOS:
                i = indices.get(campo)
                if i is not None and i < len(fila) and fila[i] is not None:
                    registro[campo] = str(fila[i]).strip()
            registro["nombre"] = limpiar_nombre(registro["nombre"])
            if registro["nombre"].lower() == "nombre" or not nombre_valido(registro["nombre"]):
                continue
            clave = registro_key(registro)
            if not clave:
                continue
            if clave in registros:
                registros[clave] = fusionar_registro(registros[clave], registro)
            else:
                registros[clave] = registro
        wb.close()
    except Exception as e:
        print("Aviso leyendo Excel anterior:", e)
    return registros


def importar_csv_viejo():
    """Recupera el CSV anterior y corrige nombres heredados."""
    registros = {}
    if not CSV_VIEJO.exists():
        return registros
    try:
        with open(CSV_VIEJO, "r", encoding="utf-8-sig", newline="") as f:
            lector = csv.DictReader(f)
            for fila in lector:
                registro = registro_vacio()
                for campo in CAMPOS:
                    registro[campo] = (fila.get(campo, "") or "").strip()
                registro["nombre"] = limpiar_nombre(registro["nombre"])
                if registro["nombre"].lower() == "nombre" or not nombre_valido(registro["nombre"]):
                    continue
                clave = registro_key(registro)
                if not clave:
                    continue
                if clave in registros:
                    registros[clave] = fusionar_registro(registros[clave], registro)
                else:
                    registros[clave] = registro
    except Exception as e:
        print("Aviso importando CSV anterior:", e)
    return registros


def guardar_excel(registros):
    """
    Guarda SIEMPRE un XLSX real.
    """

    wb = Workbook()

    ws = wb.active

    ws.title = "Marketplace"

    # Encabezados.
    for col, campo in enumerate(
        CAMPOS,
        1
    ):

        celda = ws.cell(
            row=1,
            column=col,
            value=campo
        )

        celda.font = Font(
            bold=True
        )

    fila = 2
    vistos = set()

    for registro in registros.values():
        registro["nombre"] = limpiar_nombre(registro.get("nombre", ""))
        if not nombre_valido(registro["nombre"]):
            continue
        clave = registro_key(registro)
        if not clave or clave in vistos:
            continue
        vistos.add(clave)

        for col, campo in enumerate(
            CAMPOS,
            1
        ):

            valor = registro.get(
                campo,
                ""
            )

            celda = ws.cell(
                row=fila,
                column=col
            )

            # IDs siempre como TEXTO.
            if campo in {
                "sku",
                "listing_id"
            }:

                celda.number_format = "@"

                celda.value = (
                    str(valor)
                    if valor
                    else ""
                )

            else:

                celda.value = valor

        fila += 1

    ws.freeze_panes = "A2"

    ws.auto_filter.ref = (
        ws.dimensions
    )

    anchos = {
        "A": 70,
        "B": 14,
        "C": 18,
        "D": 18,
        "E": 28,
        "F": 25,
    }

    for columna, ancho in anchos.items():

        ws.column_dimensions[
            columna
        ].width = ancho

    wb.save(
        EXCEL_SALIDA
    )


# ============================================================
# DETECTAR TARJETAS
# ============================================================

def detectar_tarjetas(page):

    """
    Mantiene el enfoque del código original:
    buscar img[alt] y subir por los padres.

    Correcciones:
    - El nombre sale del ALT de la imagen.
    - No dependemos de "Tu publicación".
    - Se intenta primero un contenedor role=button.
    - Después cualquier ancestro razonable con precio.
    """

    resultado = []

    try:

        imagenes = page.locator(
            "img[alt]"
        )

        cantidad = imagenes.count()

        print(
            "Imágenes encontradas:",
            cantidad
        )

        for i in range(
            cantidad
        ):

            try:

                img = imagenes.nth(i)

                alt = (
                    img.get_attribute(
                        "alt"
                    )
                    or ""
                ).strip()

                if not nombre_valido(
                    alt
                ):
                    continue

                alt_lower = alt.lower()

                if any(
                    x in alt_lower
                    for x in (
                        "facebook",
                        "marketplace",
                        "avatar",
                        "perfil",
                        "logo",
                        "icon",
                    )
                ):
                    continue

                candidato = None
                texto = ""

                # ------------------------------------
                # 1. Ancestro role=button.
                # ------------------------------------

                for nivel in range(
                    1,
                    11
                ):

                    try:

                        padre = img.locator(
                            "xpath="
                            + "/.." * nivel
                        )

                        if padre.count() == 0:
                            continue

                        role = (
                            padre.get_attribute(
                                "role"
                            )
                            or ""
                        )

                        if role.lower() != "button":
                            continue

                        t = limpiar_texto(
                            padre.inner_text(
                                timeout=800
                            )
                        )

                        if not t:
                            continue

                        if not re.search(
                            r"S/\s*[\d.,]+",
                            t,
                            re.I
                        ):
                            continue

                        if len(t) > 1800:
                            continue

                        candidato = padre
                        texto = t

                        break

                    except Exception:
                        pass

                # ------------------------------------
                # 2. Fallback: ancestro con precio.
                # ------------------------------------

                if candidato is None:

                    for nivel in range(
                        2,
                        11
                    ):

                        try:

                            padre = img.locator(
                                "xpath="
                                + "/.." * nivel
                            )

                            if padre.count() == 0:
                                continue

                            t = limpiar_texto(
                                padre.inner_text(
                                    timeout=800
                                )
                            )

                            if not t:
                                continue

                            if not re.search(
                                r"S/\s*[\d.,]+",
                                t,
                                re.I
                            ):
                                continue

                            if len(t) > 1200:
                                continue

                            candidato = padre
                            texto = t

                            break

                        except Exception:
                            pass

                if candidato is None:
                    continue

                # ------------------------------------
                # NOMBRE: ALT tiene prioridad absoluta.
                # ------------------------------------

                nombre = nombre_desde_tarjeta(
                    img,
                    texto
                )

                if not nombre_valido(
                    nombre
                ):
                    continue

                p = precio(
                    texto
                )

                if not p:
                    continue

                resultado.append({
                    "locator": candidato,
                    "img": img,
                    "nombre": nombre,
                    "precio": p,
                    "disponibilidad": disponibilidad(
                        texto
                    ),
                })

            except Exception:
                pass

    except Exception as e:

        print(
            "Error detectando tarjetas:",
            e
        )

    return resultado


# ============================================================
# CARGAR TODO
# ============================================================

def cargar_todo(page):

    print(
        "\nCargando publicaciones...\n"
    )

    ultimo = -1
    sin_cambios = 0

    for vuelta in range(
        1,
        MAX_SCROLLS + 1
    ):

        try:

            actual = page.locator(
                "img[alt]"
            ).count()

        except Exception:

            actual = 0

        print(
            f"Scroll {vuelta:02d} | "
            f"imágenes DOM: {actual}"
        )

        if actual == ultimo:

            sin_cambios += 1

        else:

            sin_cambios = 0
            ultimo = actual

        if sin_cambios >= 5:
            break

        try:

            # Primero scroll normal.
            page.mouse.wheel(
                0,
                1300
            )

        except Exception:

            break

        time.sleep(
            PAUSA_SCROLL
        )

    try:

        page.evaluate(
            "window.scrollTo(0,0)"
        )

    except Exception:
        pass

    time.sleep(
        0.5
    )


# ============================================================
# MODAL
# ============================================================

def obtener_modal(page):

    selectores = [
        '[role="dialog"][aria-label="Tu publicación"]',
        '[aria-label="Tu publicación"][role="dialog"]',
    ]

    for selector in selectores:

        try:

            loc = page.locator(
                selector
            )

            if loc.count():

                modal = loc.last

                if modal.is_visible(
                    timeout=500
                ):
                    return modal

        except Exception:
            pass

    try:

        loc = page.locator(
            '[role="dialog"]'
        )

        if loc.count():
            return loc.last

    except Exception:
        pass

    return None


def cerrar_modal(page):

    try:

        page.keyboard.press(
            "Escape"
        )

        time.sleep(
            0.2
        )

    except Exception:
        pass


# ============================================================
# EXTRAER PUBLICACIÓN
# ============================================================

def extraer_publicacion(
    page,
    info
):

    # ESTE ES EL NOMBRE DE LA TARJETA.
    # Nunca se reemplaza por "Tu publicación".
    nombre_tarjeta = info[
        "nombre"
    ]

    precio_tarjeta = info[
        "precio"
    ]

    print(
        "\n--------------------------------------------"
    )

    print(
        "Abriendo:",
        nombre_tarjeta
    )

    try:

        info[
            "locator"
        ].scroll_into_view_if_needed(
            timeout=3000
        )

        time.sleep(
            0.1
        )

        info[
            "locator"
        ].click(
            timeout=3000
        )

    except Exception:

        try:

            info[
                "img"
            ].click(
                timeout=3000
            )

        except Exception as e:

            print(
                "No se pudo abrir:",
                e
            )

            return {
                "nombre": nombre_tarjeta,
                "precio": precio_tarjeta,
                "disponibilidad": info.get(
                    "disponibilidad",
                    ""
                ),
                "sku": "",
                "publicado": "",
                "listing_id": "",
            }

    try:

        modal = page.locator(
            '[role="dialog"][aria-label="Tu publicación"]'
        ).last

        modal.wait_for(
            state="visible",
            timeout=TIMEOUT_MODAL
        )

    except Exception:

        modal = obtener_modal(
            page
        )

    if not modal:

        print(
            "No apareció el modal."
        )

        cerrar_modal(
            page
        )

        return {
            "nombre": nombre_tarjeta,
            "precio": precio_tarjeta,
            "disponibilidad": info.get(
                "disponibilidad",
                ""
            ),
            "sku": "",
            "publicado": "",
            "listing_id": "",
        }

    try:

        texto = limpiar_texto(
            modal.inner_text(
                timeout=2000
            )
        )

    except Exception:

        texto = ""

    # --------------------------------------------------------
    # IMPORTANTE:
    # NOMBRE SIEMPRE VIENE DE LA TARJETA.
    # NO usamos nombre_desde_texto(texto) aquí.
    # Esto elimina definitivamente el error "Tu publicación".
    # --------------------------------------------------------

    s = sku_desde_modal(
        texto
    )

    ps = precios(
        texto
    )

    estado = disponibilidad(
        texto
    )

    fecha = publicado(
        texto
    )

    listing = listing_id_modal(
        modal
    )

    registro = {
        "nombre": nombre_tarjeta,
        "precio": ps[0]
        if ps
        else precio_tarjeta,
        "disponibilidad": estado
        or info.get(
            "disponibilidad",
            ""
        ),
        "sku": s,
        "publicado": fecha,
        "listing_id": listing,
    }

    print(
        json.dumps(
            registro,
            ensure_ascii=False,
            indent=2
        )
    )

    cerrar_modal(
        page
    )

    time.sleep(
        PAUSA_ENTRE_PRODUCTOS
    )

    return registro


# ============================================================
# MAIN
# ============================================================

def main():
    print("\n============================================")
    print(" EXTRACTOR FACEBOOK MARKETPLACE")
    print("============================================\n")

    registros = cargar_excel()

    if not registros:
        recuperados = importar_csv_viejo()
        if recuperados:
            registros.update(recuperados)
            guardar_excel(registros)
            print("Registros recuperados del CSV anterior:", len(registros))

    print("Registros existentes:", len(registros))

    ids_existentes = {
        str(x.get("listing_id") or "").strip()
        for x in registros.values()
        if str(x.get("listing_id") or "").strip()
    }

    with sync_playwright() as p:
        print("Conectando al Chrome existente...")
        try:
            browser = p.chromium.connect_over_cdp(CDP_URL)
        except Exception as e:
            print("\nNO SE PUDO CONECTAR AL CHROME.")
            print("Inicia Chrome con --remote-debugging-port=9222")
            print(e)
            input("\nENTER para salir...")
            return

        print("Chrome conectado correctamente.")
        page = None

        # Buscar primero la pestaña de Marketplace.
        for context in browser.contexts:
            for pestaña in context.pages:
                try:
                    url = (pestaña.url or "").lower()
                    if "facebook.com/marketplace" in url:
                        page = pestaña
                        if "marketplace/you/selling" in url:
                            break
                except Exception:
                    pass
            if page and "marketplace/you/selling" in page.url.lower():
                break

        if page is None:
            for context in browser.contexts:
                if context.pages:
                    page = context.pages[0]
                    break

        if page is None:
            print("No hay pestañas disponibles.")
            input("\nENTER para salir...")
            return

        print("Página seleccionada:")
        print(page.url)

        if "marketplace/you/selling" not in page.url.lower():
            print("Abriendo Marketplace...")
            try:
                page.goto(MARKETPLACE_URL, wait_until="domcontentloaded", timeout=30000)
            except Exception as e:
                print("Aviso:", e)

        print("Esperando Marketplace...")
        page.wait_for_timeout(1200)
        print("URL actual:")
        print(page.url)

        cargar_todo(page)
        tarjetas = detectar_tarjetas(page)

        print("\n============================================")
        print("TARJETAS DETECTADAS:", len(tarjetas))
        print("============================================\n")

        for tarjeta in tarjetas[:10]:
            print("  ->", tarjeta["nombre"], "|", "S/" + tarjeta["precio"])

        if not tarjetas:
            print("No se detectaron tarjetas.")
            print("No se modifica el Excel existente.")
            input("\nENTER para salir...")
            return

        procesadas = 0
        saltadas = 0
        corregidas = 0

        # IMPORTANTE: procesamos todas las tarjetas.
        # No saltamos por nombre; el listing_id decide si es la misma publicación.
        for i, info in enumerate(tarjetas, 1):
            nombre = limpiar_nombre(info["nombre"])
            print(f"\n[{i}/{len(tarjetas)}]")

            registro = extraer_publicacion(page, info)
            registro["nombre"] = limpiar_nombre(registro.get("nombre") or nombre)
            if not nombre_valido(registro["nombre"]):
                registro["nombre"] = nombre

            listing = str(registro.get("listing_id") or "").strip()

            # MISMO listing_id => actualizar, jamás duplicar.
            if listing and listing in ids_existentes:
                clave = "id:" + listing
                if clave in registros:
                    registros[clave] = fusionar_registro(registros[clave], registro)
                else:
                    registros[clave] = registro
                print("Misma publicación: datos actualizados, NO se duplica.")
                corregidas += 1
                saltadas += 1
                guardar_excel(registros)
                continue

            if listing:
                # Si existe una fila antigua sin listing_id y mismo nombre,
                # la completamos; no creamos otra fila.
                nombre_norm = normalizar(registro["nombre"])
                antigua_sin_id = None
                for k, v in list(registros.items()):
                    if (not str(v.get("listing_id") or "").strip()
                            and normalizar(v.get("nombre", "")) == nombre_norm):
                        antigua_sin_id = k
                        break

                if antigua_sin_id:
                    registro = fusionar_registro(registros.pop(antigua_sin_id), registro)
                    print("Registro anterior sin ID: completado con listing_id.")

                registros["id:" + listing] = registro
                ids_existentes.add(listing)
            else:
                # Sin listing_id no podemos afirmar duplicado. Solo consolidamos
                # con una fila que tampoco tenga ID y tenga el mismo nombre.
                clave_nombre = "name:" + normalizar(registro["nombre"])
                if clave_nombre in registros:
                    registros[clave_nombre] = fusionar_registro(registros[clave_nombre], registro)
                    print("Publicación sin listing_id: fila existente completada.")
                else:
                    registros[clave_nombre] = registro
                    print("Nueva publicación sin listing_id.")

            guardar_excel(registros)
            procesadas += 1

        guardar_excel(registros)

        print("\n============================================")
        print(" EXTRACCIÓN TERMINADA")
        print("============================================")
        print("Tarjetas:", len(tarjetas))
        print("Procesadas:", procesadas)
        print("Actualizadas sin duplicar:", corregidas)
        print("Saltadas por listing_id:", saltadas)
        print("Total Excel:", len(registros))
        print("\nEXCEL:")
        print(EXCEL_SALIDA)

        total = len(registros)
        sku_ok = sum(bool(x.get("sku")) for x in registros.values())
        id_ok = sum(bool(x.get("listing_id")) for x in registros.values())
        estado_ok = sum(bool(x.get("disponibilidad")) for x in registros.values())
        publicado_ok = sum(bool(x.get("publicado")) for x in registros.values())

        print("\nCALIDAD")
        print(f"SKU:            {sku_ok}/{total}")
        print(f"Listing ID:     {id_ok}/{total}")
        print(f"Disponibilidad: {estado_ok}/{total}")
        print(f"Publicado:      {publicado_ok}/{total}")
        input("\nENTER para salir...")


if __name__ == "__main__":
    main()